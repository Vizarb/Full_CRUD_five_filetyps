import csv
import json
from openpyxl import Workbook, load_workbook
import sqlite3
from lxml import etree

# 1. CSV (Comma-Separated Values)

# Create and Write CSV Data
def create_csv(filename, data):
    with open(filename, 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerows(data)
    print(f"CSV file '{filename}' has been created.")

# Read CSV Data
def read_csv(filename):
    with open(filename, 'r') as file:
        reader = csv.reader(file)
        for row in reader:
            print(row)

# Update CSV Data
def update_csv(filename, old_data, new_data):
    # Read all data
    with open(filename, 'r') as file:
        rows = list(csv.reader(file))
    
    # Update the data
    for old_row, new_row in zip(old_data, new_data):
        for i, row in enumerate(rows):
            if row == old_row:
                rows[i] = new_row
    
    # Write updated data
    with open(filename, 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerows(rows)
    print(f"CSV file '{filename}' has been updated.")

# Delete CSV Data
def delete_csv(filename, data_to_delete):
    # Read all data
    with open(filename, 'r') as file:
        rows = list(csv.reader(file))
    
    # Remove the data
    rows = [row for row in rows if row not in data_to_delete]
    
    # Write updated data
    with open(filename, 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerows(rows)
    print(f"CSV file '{filename}' has been updated to remove specified data.")

# Example CSV data
csv_data = [
    ['Name', 'Age', 'Occupation'],
    ['Alice', 30, 'Engineer'],
    ['Bob', 25, 'Data Scientist'],
    ['Charlie', 35, 'Teacher']
]

csv_filename = 'data.csv'
create_csv(csv_filename, csv_data)
print("Reading CSV file...")
read_csv(csv_filename)
update_csv(csv_filename, [['Bob', 25, 'Data Scientist']], [['Bob', 26, 'Senior Data Scientist']])
print("Reading updated CSV file...")
read_csv(csv_filename)
delete_csv(csv_filename, [['Charlie', 35]])
print("Reading CSV file after deletion...")
read_csv(csv_filename)

# 2. JSON (JavaScript Object Notation)

# Create and Write JSON Data
def create_json(filename, data):
    with open(filename, 'w') as file:
        json.dump(data, file, indent=4)
    print(f"JSON file '{filename}' has been created.")

# Read JSON Data
def read_json(filename):
    with open(filename, 'r') as file:
        data = json.load(file)
    print(data)

# Update JSON Data
def update_json(filename, old_data, new_data):
    with open(filename, 'r') as file:
        data = json.load(file)
    
    # Update data (assuming single-level dictionary for simplicity)
    if data == old_data:
        data = new_data
    
    with open(filename, 'w') as file:
        json.dump(data, file, indent=4)
    print(f"JSON file '{filename}' has been updated.")

# Delete JSON Data
def delete_json(filename, data_to_delete):
    with open(filename, 'r') as file:
        data = json.load(file)
    
    # Delete data (assuming single-level dictionary for simplicity)
    if data == data_to_delete:
        data = {}
    
    with open(filename, 'w') as file:
        json.dump(data, file, indent=4)
    print(f"JSON file '{filename}' has been updated to remove specified data.")

# Example JSON data
json_data = {
    'Name': 'Alice',
    'Age': 30,
    'Occupation': 'Engineer'
}

json_filename = 'data.json'
create_json(json_filename, json_data)
print("Reading JSON file...")
read_json(json_filename)
update_json(json_filename, json_data, {'Name': 'Alice', 'Age': 31, 'Occupation': 'Senior Engineer'})
print("Reading updated JSON file...")
read_json(json_filename)
delete_json(json_filename, {'Name': 'Alice', 'Age': 31, 'Occupation': 'Senior Engineer'})
print("Reading JSON file after deletion...")
read_json(json_filename)

# 3. Excel

# Create and Write Excel Data
def create_excel(filename, data):
    wb = Workbook()
    ws = wb.active
    
    # Write header
    ws.append(data[0])
    
    # Write rows
    for row in data[1:]:
        ws.append(row)
    
    wb.save(filename)
    print(f"Excel file '{filename}' has been created.")

# Read Excel Data
def read_excel(filename):
    wb = load_workbook(filename)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        print(row)

# Update Excel Data
def update_excel(filename, old_data, new_data):
    wb = load_workbook(filename)
    ws = wb.active
    
    # Convert rows to lists for easier manipulation
    rows = list(ws.iter_rows(values_only=True))
    
    # Update data
    for old_row, new_row in zip(old_data, new_data):
        for i, row in enumerate(rows):
            if list(row) == old_row:
                rows[i] = tuple(new_row)
    
    # Write updated data
    ws.delete_rows(1, ws.max_row)
    for row in rows:
        ws.append(row)
    
    wb.save(filename)
    print(f"Excel file '{filename}' has been updated.")

# Delete Excel Data
def delete_excel(filename, data_to_delete):
    wb = load_workbook(filename)
    ws = wb.active
    
    # Convert rows to lists for easier manipulation
    rows = list(ws.iter_rows(values_only=True))
    
    # Remove data
    rows = [row for row in rows if row not in data_to_delete]
    
    # Write updated data
    ws.delete_rows(1, ws.max_row)
    for row in rows:
        ws.append(row)
    
    wb.save(filename)
    print(f"Excel file '{filename}' has been updated to remove specified data.")

# Example Excel data
excel_data = [
    ['Name', 'Age', 'Occupation'],
    ['Alice', 30, 'Engineer'],
    ['Bob', 25, 'Data Scientist'],
    ['Charlie', 35, 'Teacher']
]

excel_filename = 'data.xlsx'
create_excel(excel_filename, excel_data)
print("Reading Excel file...")
read_excel(excel_filename)
update_excel(excel_filename, [['Bob', 25, 'Data Scientist']], [['Bob', 26, 'Senior Data Scientist']])
print("Reading updated Excel file...")
read_excel(excel_filename)
delete_excel(excel_filename, [['Charlie', 35]])
print("Reading Excel file after deletion...")
read_excel(excel_filename)

# 4. XML (eXtensible Markup Language)

# Create and Write XML Data
def create_xml(filename, data):
    root = etree.Element('root')
    for person in data['person']:
        person_elem = etree.SubElement(root, 'person')
        for key, value in person.items():
            child = etree.SubElement(person_elem, key)
            child.text = str(value)

    tree = etree.ElementTree(root)
    tree.write(filename, encoding='utf-8', xml_declaration=True)
    print(f"XML file '{filename}' has been created.")

# Read XML Data
def read_xml(filename):
    tree = etree.parse(filename)
    root = tree.getroot()
    for person in root.findall('person'):
        name = person.find('Name').text
        age = person.find('Age').text
        occupation = person.find('Occupation').text
        print(f"Name: {name}, Age: {age}, Occupation: {occupation}")

# Update XML Data
def update_xml(filename, old_data, new_data):
    tree = etree.parse(filename)
    root = tree.getroot()
    for old_person, new_person in zip(old_data, new_data):
        for person_elem in root.findall('person'):
            name = person_elem.find('Name').text
            if name == old_person['Name']:
                for key, value in new_person.items():
                    person_elem.find(key).text = str(value)
    tree.write(filename, encoding='utf-8', xml_declaration=True)
    print(f"XML file '{filename}' has been updated.")

# Delete XML Data
def delete_xml(filename, data_to_delete):
    tree = etree.parse(filename)
    root = tree.getroot()
    for item in data_to_delete:
        for person_elem in root.findall('person'):
            name = person_elem.find('Name').text
            if name == item['Name']:
                root.remove(person_elem)
    tree.write(filename, encoding='utf-8', xml_declaration=True)
    print(f"XML file '{filename}' has been updated to remove specified data.")

# Example XML data
xml_data = {
    'person': [
        {'Name': 'Alice', 'Age': 30, 'Occupation': 'Engineer'},
        {'Name': 'Bob', 'Age': 25, 'Occupation': 'Data Scientist'},
        {'Name': 'Charlie', 'Age': 35, 'Occupation': 'Teacher'}
    ]
}

xml_filename = 'data.xml'
create_xml(xml_filename, xml_data)
print("Reading XML file...")
read_xml(xml_filename)
update_xml(xml_filename, [{'Name': 'Bob'}], [{'Name': 'Bob', 'Age': 26, 'Occupation': 'Senior Data Scientist'}])
print("Reading updated XML file...")
read_xml(xml_filename)
delete_xml(xml_filename, [{'Name': 'Charlie'}])
print("Reading XML file after deletion...")
read_xml(xml_filename)

# 5. SQLite

# Create and Insert Data into SQLite Database
def create_sqlite(filename):
    conn = sqlite3.connect(filename)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS people (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            age INTEGER,
            occupation TEXT
        )
    ''')
    conn.commit()
    conn.close()
    print(f"SQLite database '{filename}' has been created.")

def insert_sqlite(filename, data):
    conn = sqlite3.connect(filename)
    cursor = conn.cursor()
    cursor.executemany('INSERT INTO people (name, age, occupation) VALUES (?, ?, ?)', data)
    conn.commit()
    conn.close()
    print("Data has been inserted into SQLite database.")

# Read Data from SQLite Database
def read_sqlite(filename):
    conn = sqlite3.connect(filename)
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM people')
    rows = cursor.fetchall()
    for row in rows:
        print(row)
    conn.close()

# Update Data in SQLite Database
def update_sqlite(filename, old_data, new_data):
    conn = sqlite3.connect(filename)
    cursor = conn.cursor()
    for old_entry, new_entry in zip(old_data, new_data):
        cursor.execute('UPDATE people SET name = ?, age = ?, occupation = ? WHERE name = ? AND age = ?',
                       (new_entry[0], new_entry[1], new_entry[2], old_entry[0], old_entry[1]))
    conn.commit()
    conn.close()
    print("SQLite database has been updated.")

# Delete Data from SQLite Database
def delete_sqlite(filename, data_to_delete):
    conn = sqlite3.connect(filename)
    cursor = conn.cursor()
    for item in data_to_delete:
        cursor.execute('DELETE FROM people WHERE name = ? AND age = ?', (item[0], item[1]))
    conn.commit()
    conn.close()
    print("Data has been deleted from SQLite database.")

# Example SQLite data
sqlite_filename = 'data.db'
create_sqlite(sqlite_filename)
insert_sqlite(sqlite_filename, [('Alice', 30, 'Engineer'), ('Bob', 25, 'Data Scientist'), ('Charlie', 35, 'Teacher')])
print("Reading SQLite database...")
read_sqlite(sqlite_filename)
update_sqlite(sqlite_filename, [('Bob', 25)], [('Bob', 26, 'Senior Data Scientist')])
print("Reading updated SQLite database...")
read_sqlite(sqlite_filename)
delete_sqlite(sqlite_filename, [('Charlie', 35)])
print("Reading SQLite database after deletion...")
read_sqlite(sqlite_filename)
